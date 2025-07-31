import openpyxl
import unicodedata
import re
import os

# Normalizar texto
def normalize_text(texto):
    if texto is None:
        return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = re.sub(r'[\u0300-\u036f]', '', texto)
    texto = re.sub(r'[^a-z\s]', '', texto)
    return texto.strip()

# Contar periferias
def extraer_periferias(texto):
    if texto is None:
        return 0
    texto = texto.lower()
    return texto.count("periferia")

# Leer archivo tarifario
def leer_tarifario(hoja_tarifario):
    datos = []
    for fila in range(4, hoja_tarifario.max_row + 1):
        origen = hoja_tarifario[f'B{fila}'].value
        destino = hoja_tarifario[f'C{fila}'].value
        if origen and destino:
            datos.append((fila, origen, destino))
    return datos

# Obtener candidatos
def obtener_candidatos(hoja_ubicaciones, texto):
    texto_norm = normalize_text(texto)
    palabras = texto_norm.split()
    # Diccionario de departamentos y códigos Divipola
    departamentos_codigos = {
        "amazonas": ["91"], "antioquia": ["5"], "arauca": ["81"], "atlantico": ["8"],
        "bogota": ["11"], "bolivar": ["13"], "boyaca": ["15"], "caldas": ["17"],
        "caqueta": ["18"], "casanare": ["85"], "cauca": ["19", "76"], "cesar": ["20"],
        "choco": ["27"], "cordoba": ["23"], "cundinamarca": ["25"], "guainia": ["94"],
        "guaviare": ["95"], "huila": ["41"], "guajira": ["44"], "magdalena": ["47"],
        "meta": ["50"], "narino": ["52"], "putumayo": ["86"], "quindio": ["63"],
        "risaralda": ["66"], "santander": ["68", "54"], "sucre": ["70"], "tolima": ["73"],
        "vaupes": ["97"], "vichada": ["99"], "valle": ["76"], "norte": ["54"]
    }
    municipios_con_departamento_en_nombre = {
        "san jacinto del cauca",
        "puerto santander",
        "san juan del cesar",
        "santander de quilichao",
        "risaralda",
        "cordoba",
        "nariño",
        "sucre",
        "bolivar",
        "caldas"
    }

    def detectar_municipio_en_texto(t, municipios_con_departamento_en_nombre):
        palabras = t.split()  # Divide en palabras
        municipios_ordenados = sorted(
            municipios_con_departamento_en_nombre, 
            key=lambda m: len(m.split()), 
            reverse=True
        )
        for municipio in municipios_ordenados:
            num_palabras_mun = len(municipio.split())
            if len(palabras) >= num_palabras_mun:
                texto_a_comparar = " ".join(palabras[:num_palabras_mun])
                if texto_a_comparar == municipio:
                    return municipio
        return None
    def municipio_sin_dep():
        municipio_bus = texto_norm
        for fila in range(2, hoja_ubicaciones.max_row + 1):
            municipio = normalize_text(str(hoja_ubicaciones[f'D{fila}'].value))
            if municipio_bus == municipio:
                depto = hoja_ubicaciones[f'B{fila}'].value
                cod_divipola = str(hoja_ubicaciones[f'C{fila}'].value) + "000"
                candidatos.append((fila, cod_divipola, depto, municipio))
        print("Búsqueda por municipio directo:", candidatos)
        return candidatos
# Normalizar nombres de departamento
    a = detectar_municipio_en_texto(texto_norm, municipios_con_departamento_en_nombre)
    candidatos = []

    if a:
        texto_restante = texto_norm[len(a):].strip()
        palabras_restantes = texto_restante.split()
        departamento_detectado = None

        for palabra in palabras_restantes:
            if palabra.lower() in departamentos_codigos:
                departamento_detectado = palabra.lower()
                break  # Nos quedamos con el primer departamento válido encontrado

        if not departamento_detectado:
            print(f"Municipio '{a}' detectado pero sin departamento válido después. Se omite.")
            return municipio_sin_dep()

        codigos_depto = departamentos_codigos[departamento_detectado]
        municipio_busqueda = normalize_text(a)  # ya detectado
        
        for fila in range(2, hoja_ubicaciones.max_row + 1):
            cod_depto = str(hoja_ubicaciones[f'A{fila}'].value)
            if cod_depto in codigos_depto:
                municipio = normalize_text(str(hoja_ubicaciones[f'D{fila}'].value))
                if municipio_busqueda == municipio:
                    depto = hoja_ubicaciones[f'B{fila}'].value
                    cod_divipola = str(hoja_ubicaciones[f'C{fila}'].value) + "000"
                    candidatos.append((fila, cod_divipola, depto, municipio))

                    print("Búsqueda por municipio PROBLEMATICO con departamento detectado:", candidatos)
                    return candidatos

    # Si el texto incluye un municipio de los problemáticos, no hagas búsqueda por departamento
    elif len(palabras) == 1:
        municipio_busqueda = texto_norm
        if municipio_busqueda == "bogota":
            municipio_busqueda = "bogota dc"  # Normalizar Bogotá DC
        for fila in range(2, hoja_ubicaciones.max_row + 1):
            municipio = normalize_text(str(hoja_ubicaciones[f'D{fila}'].value))
            if municipio_busqueda == municipio:
                depto = hoja_ubicaciones[f'B{fila}'].value
                cod_divipola = str(hoja_ubicaciones[f'C{fila}'].value) + "000"
                candidatos.append((fila, cod_divipola, depto, municipio))
        print("Búsqueda por municipio directo:", candidatos)
        return candidatos
    else:
        # Si hay más de una palabra, buscar departamento desde la segunda palabra en adelante
        codigo_departamento = None
        for i in range(1, len(palabras)):
            palabra = palabras[i]
            if palabra in departamentos_codigos:
                if palabra == "bogota":
                    print("Encontrado Bogotá DC")
                    if i + 1 < len(palabras) and palabras[i + 1] in ["dc", "d", "d.c."]:
                        codigo_departamento = departamentos_codigos[palabra]
                        break
                elif palabra == "caldas":
                    if i >= 2 and palabras[i - 2] == "valle" and palabras[i - 1] == "del":
                        codigo_departamento = ["76"]  # Valle del Cauca
                    else:
                        codigo_departamento = ["17"]  # Caldas
                    break

                elif palabra == "santander":
                    if i >= 2 and palabras[i - 2] == "norte" and palabras[i - 1] == "de":
                        codigo_departamento = ["54"]  # Norte de Santander
                    else:
                        codigo_departamento = ["68"]  # Santander
                    break
                else:
                    codigo_departamento = departamentos_codigos[palabra]
                    break
        print("Código departamento encontrado:", codigo_departamento)
        # Si encontró un departamento, buscar municipios solo en ese departamento
        if "11" in codigo_departamento:
            municipio = hoja_ubicaciones[f'D{150}'].value
            depto = hoja_ubicaciones[f'B{150}'].value
            cod_divipola = str(hoja_ubicaciones[f'C{150}'].value) + "000"
            candidatos.append((150, cod_divipola, depto, municipio))
            print("Bogota DC encontrado:", candidatos)
            return candidatos
        elif codigo_departamento:
            for fila in range(2, hoja_ubicaciones.max_row + 1):
                cod_depto = str(hoja_ubicaciones[f'A{fila}'].value)
                if cod_depto in codigo_departamento:
                    d = normalize_text(str(hoja_ubicaciones[f'B{fila}'].value))
                    municipio_busqueda = texto_norm.replace(d, "").strip()
                    municipio = normalize_text(hoja_ubicaciones[f'D{fila}'].value)
                    if municipio_busqueda == municipio:
                        municipio = hoja_ubicaciones[f'D{fila}'].value
                        depto = hoja_ubicaciones[f'B{fila}'].value
                        cod_divipola = str(hoja_ubicaciones[f'C{fila}'].value) + "000"
                        candidatos.append((fila, cod_divipola, depto, municipio))
            print("Búsqueda por departamento:", candidatos)
            return candidatos
        else: 
            # Si no encontró departamento, buscar como municipio en todo
            municipio_sin_dep()

    print("Búsqueda general por municipio (sin depto):", candidatos)
    return candidatos

# Búsqueda y actualización
def buscar_en_maestro_con_ubicaciones(hoja_tarifario, hoja_maestro, hoja_ubicaciones, datos, tipo_carga, unidad_transporte, horas_logisticas):
    try:
        tipo_carga = int(tipo_carga)
    except ValueError:
        raise ValueError("El tipo de carga debe ser un número entero (mes).")

    try:
        horas_logisticas = int(horas_logisticas)
    except ValueError:
        raise ValueError("Las horas logísticas deben ser un número.")
    offset = 0

    filas_validas_maestro = []
    for fila in range(2, hoja_maestro.max_row + 1):
        t_carga = hoja_maestro[f'H{fila}'].value
        tipo = normalize_text(hoja_maestro[f'K{fila}'].value)
        if t_carga == tipo_carga and tipo == normalize_text(unidad_transporte):
            cod_o_m = str(hoja_maestro[f'C{fila}'].value or "").strip()
            cod_d_m = str(hoja_maestro[f'E{fila}'].value or "").strip()
            vb = hoja_maestro[f'N{fila}'].value or 0
            va = hoja_maestro[f'O{fila}'].value or 0
            filas_validas_maestro.append((cod_o_m, cod_d_m, vb, va))

    for fila_tarifario, origen_txt, destino_txt in datos:
        fila_tarifario += offset

        origenes = obtener_candidatos(hoja_ubicaciones, origen_txt)
        destinos = obtener_candidatos(hoja_ubicaciones, destino_txt)

        if not origenes or not destinos:
            hoja_tarifario[f'E{fila_tarifario}'] = "No encontrado"
            continue

        primera = True
        combinaciones_usadas = set()

        for _, cod_ori, dep_ori, mun_ori in origenes:
            for _, cod_dest, dep_dest, mun_dest in destinos:
                if (cod_ori, cod_dest) in combinaciones_usadas:
                    continue

                encontrado = False
                for cod_o_m, cod_d_m, vb, va in filas_validas_maestro:
                    if cod_ori == cod_o_m and cod_dest == cod_d_m:
                        total = vb + va * horas_logisticas
                        texto_origen = f"{mun_ori} ({dep_ori})"
                        texto_destino = f"{mun_dest} ({dep_dest})"

                        if primera:
                            hoja_tarifario[f'B{fila_tarifario}'] = texto_origen
                            hoja_tarifario[f'C{fila_tarifario}'] = texto_destino
                            hoja_tarifario[f'E{fila_tarifario}'] = total
                            primera = False
                        else:
                            hoja_tarifario.insert_rows(fila_tarifario + 1)
                            hoja_tarifario[f'B{fila_tarifario + 1}'] = texto_origen
                            hoja_tarifario[f'C{fila_tarifario + 1}'] = texto_destino
                            hoja_tarifario[f'E{fila_tarifario + 1}'] = total
                            offset += 1

                        encontrado = True
                        combinaciones_usadas.add((cod_ori, cod_dest))
                        break

                if not encontrado:
                    texto_origen = f"{mun_ori} ({dep_ori})"
                    texto_destino = f"{mun_dest} ({dep_dest})"
                    if primera:
                        hoja_tarifario[f'B{fila_tarifario}'] = texto_origen
                        hoja_tarifario[f'C{fila_tarifario}'] = texto_destino
                        hoja_tarifario[f'E{fila_tarifario}'] = "No encontrado"
                        primera = False
                    else:
                        hoja_tarifario.insert_rows(fila_tarifario + 1)
                        hoja_tarifario[f'B{fila_tarifario + 1}'] = texto_origen
                        hoja_tarifario[f'C{fila_tarifario + 1}'] = texto_destino
                        hoja_tarifario[f'E{fila_tarifario + 1}'] = "No encontrado"
                        offset += 1
                    
                    encontrado = True
                    break

            if not encontrado and primera:
                hoja_tarifario[f'E{fila_tarifario}'] = "No encontrado"

def ejecutar_tarificador(tipo_vehiculo, tipo_carga, unidad_transporte, archivo_tarifario, maestros, horas_logisticas):
    ruta_maestro = maestros.get(tipo_vehiculo)
    if not ruta_maestro or not os.path.exists(ruta_maestro):
        raise FileNotFoundError("Tipo de vehículo no válido o archivo no encontrado.")

    if not archivo_tarifario.lower().endswith(".xlsx"):
        raise ValueError("El archivo del tarifario debe ser un .xlsx")

    libro_tarifario = openpyxl.load_workbook(archivo_tarifario)
    hoja_tarifario = libro_tarifario.active
    libro_maestro = openpyxl.load_workbook(ruta_maestro)
    hoja_maestro = libro_maestro.active
    libro_ubicaciones = openpyxl.load_workbook("Maestro_ubicaciones.xlsx")
    hoja_ubicaciones = libro_ubicaciones.active

    datos_tarifario = leer_tarifario(hoja_tarifario)
    buscar_en_maestro_con_ubicaciones(
        hoja_tarifario,
        hoja_maestro,
        hoja_ubicaciones,
        datos_tarifario,
        tipo_carga,
        unidad_transporte,
        horas_logisticas,
    )

    hoja_tarifario['B2'] = f"VEHICULO: {tipo_vehiculo}"
    hoja_tarifario['E2'] = f"HORAS LOGISTICAS: {horas_logisticas}"

    libro_tarifario.save(archivo_tarifario)
    return True